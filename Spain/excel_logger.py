import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image  # [!!] เพิ่มการ import ที่จำเป็น
from pathlib import Path

def setup_report(path, header):
    """
    Creates a new Excel workbook if it doesn't exist and adds a header row.
    """
    if Path(path).exists():
        print(f"Report file '{path}' already exists. Appending data.")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    workbook.save(path)
    print(f"Created new report file: '{path}'")

def append_row(path, data):
    """
    Appends a new row of data to the specified Excel workbook.
    If a data item is a path to an image, it embeds the image instead.
    """
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        # --- [!!] ส่วนที่แก้ไขทั้งหมด ---
        # เราจะไม่ใช้ sheet.append() อีกต่อไป แต่จะสร้าง loop เพื่อจัดการข้อมูลเอง
        
        new_row_number = sheet.max_row + 1
        
        # วนลูปเพื่อตรวจสอบข้อมูลในแต่ละเซลล์ของแถวใหม่
        for col_number, cell_data in enumerate(data, 1):
            cell = sheet.cell(row=new_row_number, column=col_number)

            # ตรวจสอบว่าข้อมูลเป็น Path รูปภาพหรือไม่ (เช็คทั้ง .png, .jpg, .jpeg)
            if isinstance(cell_data, str) and cell_data.lower().endswith(('.png', '.jpg', '.jpeg')):
                try:
                    # ถ้าใช่ ให้ทำการแทรกรูปภาพลงในเซลล์
                    img = Image(cell_data)
                    
                    # --- ปรับขนาดรูปภาพและแถวเพื่อให้แสดงผลสวยงาม ---
                    img.height = 90   # ความสูงของรูป (pixel)
                    img.width = 160   # ความกว้างของรูป (pixel)
                    sheet.row_dimensions[new_row_number].height = 70 # ความสูงของแถว (point)
                    
                    sheet.add_image(img, cell.coordinate)

                except FileNotFoundError:
                    # กรณีหาไฟล์รูปภาพไม่เจอ ให้เขียนข้อความบอกแทน
                    cell.value = "Image not found"
            else:
                # ถ้าไม่ใช่รูปภาพ ก็ให้เขียนข้อมูลลงไปตามปกติ
                cell.value = cell_data
        
        workbook.save(path)
        workbook.close()

    except FileNotFoundError:
        print(f"Error: Report file '{path}' not found. Could not log data.")